#!/usr/bin/env python3
"""用 LibreOffice 把 xlsx 的公式算過一遍，把結果寫回檔案。

為什麼需要這一步：openpyxl 只會把公式當「字串」寫進去，不會附帶算好的值。
桌機版 Excel 開檔時會自己重算所以看得到數字，但線上預覽（Excel Online、
Google 試算表、手機、Mac 預覽、GitHub）不會算，就會整片空白。
在 CI 先算一次，交付出去的檔案在哪裡打開都有數字，公式也還在。
"""
import subprocess
import sys
import tempfile
from pathlib import Path

MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def main():
    target = Path(sys.argv[1]).resolve()
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 300
    if not target.exists():
        print(f"✗ 找不到 {target}")
        return 1

    with tempfile.TemporaryDirectory() as tmp:
        profile = Path(tmp) / "profile"
        env_arg = f"-env:UserInstallation={profile.as_uri()}"

        # 先讓 LibreOffice 建出使用者設定檔，才能把巨集塞進去
        subprocess.run(["soffice", "--headless", "--terminate_after_init", env_arg],
                       capture_output=True, timeout=timeout, check=False)
        macro_dir = profile / "user" / "basic" / "Standard"
        if not macro_dir.exists():
            print("✗ LibreOffice 沒建出設定檔，跳過重算（公式仍在，只是沒有快取值）")
            return 0
        (macro_dir / "Module1.xba").write_text(MACRO)

        before = target.stat().st_mtime_ns
        subprocess.run(
            ["soffice", "--headless", "--norestore", env_arg,
             "macro:///Standard.Module1.RecalculateAndSave", str(target)],
            capture_output=True, timeout=timeout, check=False)

        if target.stat().st_mtime_ns == before:
            print("⚠ 檔案沒被改寫，重算可能沒生效")
            return 0

    from openpyxl import load_workbook
    wb = load_workbook(target, data_only=True)
    errs, cells = 0, 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("#"):
                    errs += 1
                elif c.value is not None:
                    cells += 1
    if errs:
        print(f"✗ 重算後有 {errs} 個公式錯誤（#REF!／#DIV/0! 之類），請檢查")
    else:
        print(f"✓ 已重算：{cells} 個有值的儲存格，0 個公式錯誤")
    return 0   # 就算有錯也照樣往下走，避免整批資料因為一格錯誤而沒被 commit


if __name__ == "__main__":
    sys.exit(main())
