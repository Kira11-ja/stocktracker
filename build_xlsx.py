# -*- coding: utf-8 -*-
"""美股觀察表 v2 — 只保留使用者指定的指標，以及計算這些指標所必需的原始欄位。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as gcl, column_index_from_string as cif
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule, FormulaRule

from pathlib import Path

DATA = Path(__file__).parent / "data"
FONT = "Arial"
NROW = max(21, len(open(DATA / "tickers.csv").readlines()) + 5)
QLAST = 600

H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
IN_FONT = Font(name=FONT, color="0000FF", size=10)
FX_FONT = Font(name=FONT, color="000000", size=10)
LNK_FONT = Font(name=FONT, color="008000", size=10)
BOT_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FONT = Font(name=FONT, bold=True, color="C00000", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="808080")
SEC_FONT = Font(name=FONT, bold=True, size=11, color="1F3864")

PCT, NUM0, NUM2, MULT = "0.0%", "#,##0", "#,##0.00", "0.0"
DATE = "yyyy-mm-dd"
BPS = '#,##0;[Red]-#,##0'
PP = '0.0;[Red]-0.0'
RED, YEL, GRN = "F8696B", "FFEB84", "63BE7B"

wb = openpyxl.Workbook(); wb.remove(wb.active)

# ── 讀入管線產出的 csv ──────────────────────────────────────────
import pandas as pd

MILLIONS = ["revenue", "gross_profit", "total_equity", "shares_diluted"]


def _read(name, dates=()):
    df = pd.read_csv(DATA / name)
    for c in dates:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
    return df


q_df = _read("raw_q.csv", ["period_end"])
for c in MILLIONS:                      # 絕對金額 → 百萬，跟表頭的 ($M) 一致
    if c in q_df.columns:
        q_df[c] = pd.to_numeric(q_df[c], errors="coerce") / 1e6
est_df = _read("raw_est.csv", ["fy1_end", "as_of"])
price_df = _read("raw_price.csv", ["price_date", "next_earnings", "as_of"])
tick_df = _read("tickers.csv", ["added"])

raw_q = q_df.where(pd.notna(q_df), None).to_dict("records")
raw_est = est_df.where(pd.notna(est_df), None).to_dict("records")
raw_price = price_df.where(pd.notna(price_df), None).to_dict("records")
TICKERS = tick_df.where(pd.notna(tick_df), None).to_dict("records")


def style_header(ws, row, n, start=1, h=32):
    for c in range(start, start + n):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = H_FILL, H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = h


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ===================== Raw_Q =====================
# 13 個原始欄位。每一欄都直接支撐使用者指定的指標，沒有多餘欄位。
QCOLS = [
    ("ticker",          "ticker",           9,  None, "鍵值"),
    ("period",          "period",           12, None, "鍵值"),
    ("fy",              "fy",               7,  "0",  "鍵值"),
    ("fq",              "fq",               5,  "0",  "鍵值"),
    ("period_end",      "period_end",       12, DATE, "鍵值"),
    ("is_est",          "is_est",           7,  None, "治理"),
    ("est_source",      "est_source",       12, None, "治理"),
    ("eps_basis",       "eps_basis",        11, None, "EPS 口徑：street=市場口徑，gaap=降級"),
    ("revenue",         "revenue ($M)",     13, NUM0, "營收成長率 · 毛利率分母"),
    ("gross_profit",    "gross_profit ($M)",14, NUM0, "毛利率分子"),
    ("eps_diluted_adj", "eps_diluted_adj",  14, NUM2, "EPS · PE · PEG · EPS成長率"),
    ("shares_diluted",  "shares_diluted (M)",14, NUM0, "反推 ROE 淨利 · ROE_F 的 BVPS"),
    ("total_equity",    "total_equity ($M)",14, NUM0, "ROE 分母"),
    ("dps",             "dps (每股配息)",    13, NUM2, "殖利率"),
    ("price_at_end",    "price_at_end",     12, NUM2, "歷史 PE · PEG_T · 歷史殖利率"),
]
QL = {k: gcl(i + 1) for i, (k, _, _, _, _) in enumerate(QCOLS)}
SEQ_COL, KEY_COL = len(QCOLS) + 1, len(QCOLS) + 2
QL["seq"], QL["key"] = gcl(SEQ_COL), gcl(KEY_COL)

ws = wb.create_sheet("Raw_Q")
ws.append([h for _, h, _, _, _ in QCOLS] + ["seq", "key"])
style_header(ws, 1, len(QCOLS) + 2)
for row in raw_q:
    ws.append([row.get(k) for k, _, _, _, _ in QCOLS]
              + [row.get("seq"), row.get("key")])
nq = len(raw_q)
for r in range(2, nq + 2):
    for i, (k, _, _, fmt, _) in enumerate(QCOLS):
        c = ws.cell(row=r, column=i + 1)
        c.font, c.fill = FX_FONT, BOT_FILL
        if fmt:
            c.number_format = fmt
    for cc in (SEQ_COL, KEY_COL):
        ws.cell(row=r, column=cc).font = FX_FONT
    ws.cell(row=r, column=SEQ_COL).number_format = "0"
widths(ws, {gcl(i + 1): w for i, (_, _, w, _, _) in enumerate(QCOLS)})
ws.column_dimensions[QL["seq"]].width = 6
ws.column_dimensions[QL["key"]].width = 11
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{QL['key']}{nq + 1}"
ws.cell(row=nq + 3, column=1,
        value="⚠ 黃底＝資料管線寫入區，請勿手改。seq 與 key 是公式：is_est=\"Y\" 記為 0（當季預估），"
              "實際季由新到舊排 1、2、3…；補進新一季時整欄自動重排。").font = WARN_FONT
ws.cell(row=nq + 4, column=1, value="每一欄的用途：").font = SEC_FONT
for i, (k, h, _, _, use) in enumerate(QCOLS):
    ws.cell(row=nq + 5 + i, column=1, value=f"{gcl(i+1)}  {h}").font = Font(name=FONT, size=9)
    ws.cell(row=nq + 5 + i, column=3, value=use).font = NOTE_FONT
ws.cell(row=nq + 5 + len(QCOLS) + 1, column=1,
        value="示範資料為合成數字，非真實財報。").font = NOTE_FONT

# ===================== 共用公式工具 =====================
def RQ(k):
    return f"Raw_Q!${QL[k]}$2:${QL[k]}${QLAST}"


def SF(val, *seqargs, r=2, tk=None):
    tk = tk or f"$A{r}"
    s = f'SUMIFS({RQ(val)},{RQ("ticker")},{tk}'
    for a in seqargs:
        s += f',{RQ("seq")},{a}'
    return s + ')'


def T(val, r, tk=None):   return SF(val, '">=1"', '"<=4"', r=r, tk=tk)
def PY(val, r, tk=None):  return SF(val, '">=5"', '"<=8"', r=r, tk=tk)
def FWD(val, r, tk=None): return SF(val, '">=0"', '"<=3"', r=r, tk=tk)
def G5(val, r, tk=None):  return SF(val, '">=1"', '"<="&Config!$B$5', r=r, tk=tk)
def QQ(val, k, r, tk=None): return SF(val, str(k), r=r, tk=tk)


def NI(lo, hi, r, tk=None):
    """Adjusted 淨利＝逐季 (adjusted EPS × 稀釋股數) 相乘後加總。
    沒有任何 API 提供 non-GAAP 淨利，只能這樣反推——逐季相乘比用單一股數準。"""
    tk = tk or f"$A{r}"
    return (f'SUMPRODUCT(({RQ("ticker")}={tk})*({RQ("seq")}>={lo})*({RQ("seq")}<={hi})'
            f'*{RQ("eps_diluted_adj")}*{RQ("shares_diluted")})')


def NI_FY(fyexpr, tk):
    return (f'SUMPRODUCT(({RQ("ticker")}={tk})*({RQ("fy")}={fyexpr})*({RQ("is_est")}="N")'
            f'*{RQ("eps_diluted_adj")}*{RQ("shares_diluted")})')


def LOOK(sheet, col, r, n=100):
    return f'INDEX({sheet}!${col}$2:${col}${n},MATCH($A{r},{sheet}!$A$2:$A${n},0))'


def QLOOK(colkey, seq, r, tk=None):
    tk = tk or f"$A{r}"
    return (f'INDEX(Raw_Q!${QL[colkey]}$2:${QL[colkey]}${QLAST},'
            f'MATCH({tk}&"|"&{seq},Raw_Q!${QL["key"]}$2:${QL["key"]}${QLAST},0))')


def wrap(expr, r):
    return f'=IF($A{r}="","",IFERROR({expr},""))'


def X(name, r):
    return f'{CL[name]}{r}'


# ===================== Tickers =====================
# 欄位對齊 tickers.csv：多了「標籤」（題材，可多值，用頓號分隔）。
# 等級的下拉限制拿掉了 —— 等級與產業都改成自由文字，可以自己打新的。
# ticker / 公司 / 等級 仍在 A / B / D，Calc 分頁的參照不受影響。
ws = wb.create_sheet("Tickers")
ws.append(["ticker", "公司名稱", "產業", "等級", "標籤", "加入日期", "備註"])
style_header(ws, 1, 7)
for t in TICKERS:
    ws.append([t.get("ticker"), t.get("company"), t.get("sector"),
               t.get("tier"), t.get("tags"), t.get("added"), t.get("note")])
for r in range(2, 22):
    for c in range(1, 8):
        ws.cell(row=r, column=c).font = IN_FONT
    ws.cell(row=r, column=6).number_format = DATE
widths(ws, {"A": 10, "B": 24, "C": 13, "D": 9, "E": 18, "F": 13, "G": 34})
ws.freeze_panes = "A2"
ws.cell(row=24, column=1,
        value="↑ 藍字＝手動輸入。新增股票就在第 2~21 列往下打一列，其餘分頁全自動延伸。").font = NOTE_FONT

# ===================== Raw_Est / Raw_Price =====================
ws = wb.create_sheet("Raw_Est")
EH = [("ticker", 9, None), ("eps_f1", 10, NUM2), ("eps_f2", 10, NUM2),
      ("fy1_end", 12, DATE), ("n_analysts", 11, "0"), ("as_of", 12, DATE)]
ws.append([h for h, _, _ in EH]); style_header(ws, 1, len(EH))
for row in raw_est:
    ws.append([row.get("ticker"), row.get("eps_f1"), row.get("eps_f2"),
               row.get("fy1_end"), row.get("n_analysts"), row.get("as_of")])
for r in range(2, len(raw_est) + 2):
    for i, (_, _, fmt) in enumerate(EH):
        c = ws.cell(row=r, column=i + 1); c.font, c.fill = FX_FONT, BOT_FILL
        if fmt: c.number_format = fmt
widths(ws, {gcl(i + 1): w for i, (_, w, _) in enumerate(EH)})
ws.freeze_panes = "B2"
ws.cell(row=len(raw_est) + 3, column=1,
        value="eps_f1＝本財年共識、eps_f2＝下一財年共識，皆為 Adjusted 口徑（與 EPS_TTM 同口徑）。").font = NOTE_FONT

ws = wb.create_sheet("Raw_Price")
PH = [("ticker", 9, None), ("price", 10, NUM2), ("price_date", 12, DATE),
      ("next_earnings", 13, DATE), ("as_of", 12, DATE)]
ws.append([h for h, _, _ in PH]); style_header(ws, 1, len(PH))
for row in raw_price:
    ws.append([row.get("ticker"), row.get("price"), row.get("price_date"),
               row.get("next_earnings"), row.get("as_of")])
for r in range(2, len(raw_price) + 2):
    for i, (_, _, fmt) in enumerate(PH):
        c = ws.cell(row=r, column=i + 1); c.font, c.fill = FX_FONT, BOT_FILL
        if fmt: c.number_format = fmt
widths(ws, {gcl(i + 1): w for i, (_, w, _) in enumerate(PH)})
ws.freeze_panes = "B2"

# ===================== Config =====================
ws = wb.create_sheet("Config")
ws.sheet_view.showGridLines = False
ws["A1"] = "參數設定"; ws["A1"].font = TITLE_FONT
ws["A3"], ws["B3"], ws["C3"] = "參數", "值", "用途"
style_header(ws, 3, 3)
CFG = [("目標季數（每檔應有的實際季數）", 24, "0", "_Manifest 判斷是否需要補歷史"),
       ("毛利率五年平均取樣季數", 20, "0", "5 年平均毛利率的取樣範圍"),
       ("PEG 成長率下限", 0.05, PCT, "成長率低於此值時 PEG 顯示 N/M（PEG 在低成長時失真）"),
       ("分析師家數下限", 5, "0", "低於此值時共識預估信心不足，會出旗標"),
       ("NTM 換算基準天數", 365, "0", "EPS_NTM 依 fy1_end 距今天數在 F1 / F2 之間加權")]
for i, (n, v, f, u) in enumerate(CFG):
    r = 4 + i
    ws.cell(row=r, column=1, value=n).font = Font(name=FONT, size=10)
    c = ws.cell(row=r, column=2, value=v)
    c.font, c.number_format = IN_FONT, f
    c.fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(row=r, column=3, value=u).font = NOTE_FONT
widths(ws, {"A": 34, "B": 10, "C": 66})
ws.cell(row=11, column=1, value="黃底藍字＝可以改，改了整張表的判斷規則跟著變。").font = NOTE_FONT

# ===================== Calc =====================
CALC = []
def add(name, header, fmt, w, grp, fn):
    CALC.append(dict(name=name, header=header, fmt=fmt, w=w, grp=grp, fn=fn))

add("ticker", "ticker", None, 9, "基本", lambda r: f'=IF(Tickers!A{r}="","",Tickers!A{r})')
add("company", "公司", None, 20, "基本", lambda r: wrap(LOOK("Tickers", "B", r), r))
add("tier", "等級", None, 8, "基本", lambda r: wrap(LOOK("Tickers", "D", r), r))
add("price", "股價", NUM2, 9, "基本", lambda r: wrap(LOOK("Raw_Price", "B", r), r))
add("rev_ttm", "Rev_TTM($M)", NUM0, 12, "基本", lambda r: wrap(T("revenue", r), r))

# --- EPS（Adjusted 口徑）---
add("eps_ttm", "EPS_TTM(Adj)", NUM2, 12, "估值", lambda r: wrap(T("eps_diluted_adj", r), r))
add("eps_ttm_fwd", "EPS_TTM_fwd", NUM2, 12, "估值", lambda r: wrap(FWD("eps_diluted_adj", r), r))
add("eps_ntm", "EPS_NTM", NUM2, 10, "估值",
    lambda r: wrap(
        f'{LOOK("Raw_Est","B",r)}*MIN(1,MAX(0,({LOOK("Raw_Est","D",r)}-TODAY())/Config!$B$8))'
        f'+{LOOK("Raw_Est","C",r)}*(1-MIN(1,MAX(0,({LOOK("Raw_Est","D",r)}-TODAY())/Config!$B$8)))', r))
add("eps_ttm_yoy", "EPS_TTM_YoY", PCT, 11, "估值",
    lambda r: wrap(f'{T("eps_diluted_adj", r)}/{PY("eps_diluted_adj", r)}-1', r))

# --- 估值 ---
add("pe_ttm", "PE_TTM", MULT, 9, "估值",
    lambda r: f'=IF($A{r}="","",IFERROR(IF({X("eps_ttm",r)}<=0,"N/M",'
              f'{X("price",r)}/{X("eps_ttm",r)}),"N/M"))')
add("pe_ttm_fwd", "PE_TTM_fwd", MULT, 10, "估值",
    lambda r: f'=IF($A{r}="","",IFERROR(IF({X("eps_ttm_fwd",r)}<=0,"N/M",'
              f'{X("price",r)}/{X("eps_ttm_fwd",r)}),"N/M"))')
add("pe_ntm", "PE_NTM", MULT, 9, "估值",
    lambda r: f'=IF($A{r}="","",IFERROR(IF({X("eps_ntm",r)}<=0,"N/M",'
              f'{X("price",r)}/{X("eps_ntm",r)}),"N/M"))')
add("eps_g_f", "EPS成長_F", PCT, 10, "估值",
    lambda r: wrap(f'{LOOK("Raw_Est","C",r)}/{LOOK("Raw_Est","B",r)}-1', r))
add("peg_t", "PEG_T", NUM2, 8, "估值",
    lambda r: f'=IF($A{r}="","",IFERROR(IF(OR(NOT(ISNUMBER({X("pe_ttm",r)})),'
              f'{X("eps_ttm_yoy",r)}<Config!$B$6),"N/M",'
              f'{X("pe_ttm",r)}/({X("eps_ttm_yoy",r)}*100)),"N/M"))')
add("peg_f", "PEG_F", NUM2, 8, "估值",
    lambda r: f'=IF($A{r}="","",IFERROR(IF(OR(NOT(ISNUMBER({X("pe_ntm",r)})),'
              f'{X("eps_g_f",r)}<Config!$B$6),"N/M",'
              f'{X("pe_ntm",r)}/({X("eps_g_f",r)}*100)),"N/M"))')

# --- 殖利率 ---
add("dps_ttm", "DPS_TTM", NUM2, 10, "殖利率", lambda r: wrap(T("dps", r), r))
add("div_yield", "殖利率", PCT, 9, "殖利率",
    lambda r: f'=IF($A{r}="","",IFERROR(IF({X("dps_ttm",r)}=0,"不配息",'
              f'{X("dps_ttm",r)}/{X("price",r)}),""))')
add("dps_yoy", "配息YoY", PCT, 10, "殖利率",
    lambda r: f'=IF($A{r}="","",IFERROR(IF({PY("dps", r)}=0,"",'
              f'{T("dps", r)}/{PY("dps", r)}-1),""))')

# --- 成長（加速度只做季度層）---
add("rev_yoy1", "營收YoY_最新季", PCT, 13, "成長",
    lambda r: wrap(f'{QQ("revenue",1,r)}/{QQ("revenue",5,r)}-1', r))
add("rev_qoq1", "營收QoQ_最新季", PCT, 13, "成長",
    lambda r: wrap(f'{QQ("revenue",1,r)}/{QQ("revenue",2,r)}-1', r))
add("rev_qoq2", "營收QoQ_上一季", PCT, 13, "成長",
    lambda r: wrap(f'{QQ("revenue",2,r)}/{QQ("revenue",3,r)}-1', r))
add("rev_accel", "營收加速度(pp)", PP, 13, "成長",
    lambda r: wrap(f'({X("rev_qoq1",r)}-{X("rev_qoq2",r)})*100', r))
add("momentum", "動能判讀", None, 14, "成長",
    lambda r: f'=IF($A{r}="","",IFERROR(IF(NOT(ISNUMBER({X("rev_qoq1",r)})),"",'
              f'IF({X("rev_qoq1",r)}>0,'
              f'IF({X("rev_accel",r)}>0,"成長且加速","成長但減速"),'
              f'IF({X("rev_accel",r)}>0,"衰退但收斂","衰退且惡化"))),""))')
add("rev_accel_sa", "同期加速度(pp)", PP, 13, "成長",
    lambda r: wrap(f'({X("rev_qoq1",r)}'
                   f'-({QQ("revenue",5,r)}/{QQ("revenue",6,r)}-1))*100', r))
add("eps_yoy1", "EPS_YoY_最新季", PCT, 13, "成長",
    lambda r: wrap(f'{QQ("eps_diluted_adj",1,r)}/{QQ("eps_diluted_adj",5,r)}-1', r))

# --- 毛利率 ---
add("gm_ttm", "毛利率_TTM", PCT, 11, "毛利",
    lambda r: wrap(f'{T("gross_profit", r)}/{X("rev_ttm",r)}', r))
add("gm_5y", "毛利率_5年平均", PCT, 12, "毛利",
    lambda r: wrap(f'{G5("gross_profit", r)}/{G5("revenue", r)}', r))
add("gm_spread", "GM vs 5年平均(bps)", BPS, 15, "毛利",
    lambda r: wrap(f'({X("gm_ttm",r)}-{X("gm_5y",r)})*10000', r))
add("gm_dyoy", "ΔGM_YoY(bps)", BPS, 12, "毛利",
    lambda r: wrap(f'({QQ("gross_profit",1,r)}/{QQ("revenue",1,r)}'
                   f'-{QQ("gross_profit",5,r)}/{QQ("revenue",5,r)})*10000', r))

# --- ROE ---
add("bvps", "每股淨值", NUM2, 10, "ROE",
    lambda r: wrap(f'{QQ("total_equity",1,r)}/{QQ("shares_diluted",1,r)}', r))
add("ni_ttm", "淨利_TTM($M)", NUM0, 13, "ROE", lambda r: wrap(NI(1, 4, r), r))
add("roe_ttm", "ROE_TTM", PCT, 10, "ROE",
    lambda r: f'=IF($A{r}="","",IFERROR(IF(MIN({QQ("total_equity",1,r)},'
              f'{QQ("total_equity",5,r)})<=0,"N/M",'
              f'{X("ni_ttm",r)}/(({QQ("total_equity",1,r)}'
              f'+{QQ("total_equity",5,r)})/2)),"N/M"))')
add("roe_5avg", "ROE_5年平均", PCT, 11, "ROE",
    lambda r: wrap(f'AVERAGE({X("roe_a0",r)}:{X("roe_a4",r)})', r))
add("roe_f", "ROE_Forward", PCT, 11, "ROE",
    lambda r: f'=IF($A{r}="","",IFERROR(IF({X("bvps",r)}<=0,"N/M",'
              f'{LOOK("Raw_Est","B",r)}/{X("bvps",r)}),"N/M"))')

# --- 預估 ---
add("eps_q_est", "當季EPS預估", NUM2, 11, "預估", lambda r: wrap(QQ("eps_diluted_adj", 0, r), r))
add("est_period", "預估期別", None, 11, "預估", lambda r: wrap(QLOOK("period", 0, r), r))
add("est_src", "預估來源", None, 11, "預估", lambda r: wrap(QLOOK("est_source", 0, r), r))
add("n_an", "分析師家數", "0", 10, "預估", lambda r: wrap(LOOK("Raw_Est", "E", r), r))
add("next_er", "下次財報日", DATE, 12, "預估", lambda r: wrap(LOOK("Raw_Price", "D", r), r))
add("as_of", "資料日期", DATE, 11, "預估", lambda r: wrap(LOOK("Raw_Price", "E", r), r))

def _anchor(k):
    def fn(r):
        e0, e4 = QQ("total_equity", k, r), QQ("total_equity", k + 4, r)
        return (f'=IF($A{r}="","",IFERROR(IF(MIN({e0},{e4})<=0,"",'
                f'{NI(k, k+3, r)}/(({e0}+{e4})/2)),""))')
    return fn
for i, k in enumerate([1, 5, 9, 13, 17]):
    add(f"roe_a{i}", f"ROE錨點_seq{k}", PCT, 12, "輔助", _anchor(k))

CL = {d["name"]: gcl(i + 1) for i, d in enumerate(CALC)}

ws = wb.create_sheet("Calc")
for i, d in enumerate(CALC):
    ws.cell(row=1, column=i + 1, value=d["header"])
    ws.column_dimensions[gcl(i + 1)].width = d["w"]
style_header(ws, 1, len(CALC))
for r in range(2, NROW + 1):
    for i, d in enumerate(CALC):
        c = ws.cell(row=r, column=i + 1, value=d["fn"](r))
        c.font = LNK_FONT if i == 0 else FX_FONT
        if d["fmt"]:
            c.number_format = d["fmt"]
    if d["grp"] == "輔助":
        pass
for i, d in enumerate(CALC):
    if d["grp"] == "輔助":
        ws.column_dimensions[gcl(i + 1)].hidden = True
ws.freeze_panes = "B2"
ws.cell(row=NROW + 2, column=1,
        value="Calc 是引擎，不是給人看的。淨利_TTM 是用「逐季 adjusted EPS × 稀釋股數」反推的——"
              "沒有任何 API 提供 non-GAAP 淨利，這是唯一算得出 Adjusted 口徑 ROE 的方法。").font = NOTE_FONT

# ===================== Dashboard =====================
DASH = [
    ("基本", [("ticker", "Ticker", 10), ("company", "公司", 18),
              ("tier", "等級", 8), ("price", "股價", 10)]),
    ("估值", [("pe_ttm", "PE_TTM", 9), ("pe_ntm", "PE Forward", 11),
              ("peg_t", "PEG_T", 8), ("peg_f", "PEG_F", 8)]),
    ("殖利率", [("div_yield", "殖利率", 9), ("dps_yoy", "配息YoY", 10)]),
    ("成長", [("rev_qoq1", "營收QoQ", 10), ("rev_qoq2", "上季QoQ", 10),
              ("rev_accel", "加速度(pp)", 11), ("momentum", "動能判讀", 13),
              ("rev_accel_sa", "同期加速度(pp)", 13), ("rev_yoy1", "營收YoY", 10),
              ("eps_yoy1", "EPS_YoY", 10), ("eps_ttm_yoy", "EPS_TTM_YoY", 11)]),
    ("毛利率", [("gm_ttm", "毛利率", 10), ("gm_5y", "5年平均", 10),
               ("gm_spread", "vs5年(bps)", 11), ("gm_dyoy", "ΔGM_YoY(bps)", 12)]),
    ("ROE", [("roe_ttm", "ROE_TTM", 10), ("roe_5avg", "ROE_5年平均", 11),
             ("roe_f", "ROE Forward", 11)]),
    ("預估", [("eps_q_est", "當季EPS預估", 12), ("est_period", "預估期別", 11),
              ("est_src", "來源", 11), ("next_er", "下次財報", 11)]),
]
GRP_COLORS = {"基本": "D9E2F3", "估值": "FCE4D6", "殖利率": "E7E6E6", "成長": "E2EFDA",
              "毛利率": "FFF2CC", "ROE": "F2E3F1", "預估": "DEEBF7"}
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
flat, col = [], 1
for grp, cols in DASH:
    start = col
    for name, header, w in cols:
        d = next(x for x in CALC if x["name"] == name)
        ws.cell(row=2, column=col, value=header)
        ws.column_dimensions[gcl(col)].width = w
        flat.append((col, name, d["fmt"]))
        col += 1
    g = ws.cell(row=1, column=start, value=grp)
    if col - 1 > start:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)
    g.font = Font(name=FONT, bold=True, size=10, color="1F3864")
    g.alignment = Alignment(horizontal="center")
    for cc in range(start, col):
        ws.cell(row=1, column=cc).fill = PatternFill("solid", fgColor=GRP_COLORS[grp])
NCOL = col - 1
style_header(ws, 2, NCOL)
for r in range(3, NROW + 2):
    src = r - 1
    for c, name, fmt in flat:
        cell = ws.cell(row=r, column=c,
                       value=f'=IF(Calc!$A${src}="","",Calc!{CL[name]}${src})')
        cell.font = LNK_FONT
        if fmt:
            cell.number_format = fmt
ws.freeze_panes = "B3"
ws.auto_filter.ref = f"A2:{gcl(NCOL)}{NROW + 1}"
col = 1
for grp, cols in DASH:
    if grp != "基本":
        ws.column_dimensions.group(gcl(col), gcl(col + len(cols) - 1), outline_level=1, hidden=False)
    col += len(cols)

DC = {name: gcl(c) for c, name, _ in flat}
rng = lambda n: f"{DC[n]}3:{DC[n]}{NROW + 1}"
ws.conditional_formatting.add(rng("rev_accel_sa"),
    ColorScaleRule(start_type="num", start_value=-8, start_color=RED,
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="num", end_value=8, end_color=GRN))
ws.conditional_formatting.add(rng("rev_accel"),
    ColorScaleRule(start_type="num", start_value=-30, start_color=RED,
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="num", end_value=30, end_color=GRN))
for n in ("gm_dyoy", "gm_spread"):
    ws.conditional_formatting.add(rng(n),
        ColorScaleRule(start_type="num", start_value=-200, start_color=RED,
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="num", end_value=200, end_color=GRN))
for n in ("rev_qoq1", "rev_yoy1", "div_yield", "roe_ttm"):
    ws.conditional_formatting.add(rng(n),
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color=GRN))
for n in ("peg_t", "peg_f"):
    ws.conditional_formatting.add(rng(n),
        CellIsRule(operator="lessThan", formula=["1.5"],
                   font=Font(name=FONT, bold=True, color="006100"),
                   fill=PatternFill("solid", fgColor="C6EFCE")))
for _txt, _fill, _col in [("成長且加速", "C6EFCE", "006100"), ("成長但減速", "FFEB9C", "9C5700"),
                          ("衰退但收斂", "FFEB9C", "9C5700"), ("衰退且惡化", "FFC7CE", "9C0006")]:
    ws.conditional_formatting.add(rng("momentum"),
        CellIsRule(operator="equal", formula=[f'"{_txt}"'],
                   font=Font(name=FONT, bold=True, color=_col),
                   fill=PatternFill("solid", fgColor=_fill)))
ws.conditional_formatting.add(rng("roe_ttm"),
    CellIsRule(operator="equal", formula=['"N/M"'],
               font=Font(name=FONT, italic=True, color="808080")))
ws.conditional_formatting.add(f"A3:{gcl(NCOL)}{NROW + 1}",
    FormulaRule(formula=['$C3="池子"'], fill=PatternFill("solid", fgColor="F2F2F2")))
nt = NROW + 3
ws.cell(row=nt, column=1,
        value="加速度(pp) ＝ 最新季 QoQ − 上一季 QoQ。QoQ ＝ 本季營收 ÷ 上一季營收 − 1（季度成長率）。"
              "正值代表這一季的成長速度比上一季快。").font = NOTE_FONT
ws.cell(row=nt + 1, column=1,
        value="ROE 顯示 N/M 代表股東權益為負（大量庫藏股或累積虧損），此時 ROE 沒有比較意義。"
              "殖利率顯示「不配息」代表該檔近四季沒有配息紀錄。").font = NOTE_FONT
ws.cell(row=nt + 2, column=1, value="資料來源：yfinance / yahooquery / Alpha Vantage / SEC EDGAR").font = NOTE_FONT

# ===================== Stock_Card =====================
TSEL = 'Stock_Card!$B$2'
ws = wb.create_sheet("Stock_Card")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 12, "C": 3, "D": 18, "E": 12, "F": 3, "G": 18, "H": 12,
            "I": 12, "J": 12, "K": 12, "L": 12})

def CLK(name):
    c = CL[name]
    return f'INDEX(Calc!${c}$2:${c}$100,MATCH($B$2,Calc!$A$2:$A$100,0))'

def CK(name):
    return f'=IF($B$2="","",IFERROR({CLK(name)},""))'

ws["A1"] = "個股卡"; ws["A1"].font = TITLE_FONT
ws["A2"] = "選擇股票 →"; ws["A2"].font = Font(name=FONT, bold=True, size=11)
b2 = ws["B2"]; b2.value = "AAPL"
b2.font = Font(name=FONT, bold=True, size=13, color="0000FF")
b2.fill = PatternFill("solid", fgColor="FFFF00")
b2.alignment = Alignment(horizontal="center")
dv = DataValidation(type="list", formula1="=Tickers!$A$2:$A$100", allow_blank=False)
ws.add_data_validation(dv); dv.add("B2")
ws["D2"] = CK("company"); ws["D2"].font = Font(name=FONT, bold=True, size=12)
ws.merge_cells("D2:F2")
ws["G2"] = CK("tier"); ws["G2"].font = Font(name=FONT, size=11, color="808080")
ws["C3"] = "← 這個下拉同時驅動「年度檢視」與「季度檢視」兩頁"; ws["C3"].font = NOTE_FONT

for colL, lab, name, fmt in [("A", "股價", "price", NUM2), ("D", "分析師家數", "n_an", "0"),
                             ("G", "下次財報日", "next_er", DATE), ("I", "資料日期", "as_of", DATE)]:
    ws[f"{colL}4"] = lab
    ws[f"{colL}4"].font = Font(name=FONT, size=9, color="808080")
    c = ws[f"{gcl(cif(colL) + 1)}4"]
    c.value = CK(name); c.font = Font(name=FONT, bold=True, size=11)
    if fmt: c.number_format = fmt

BLOCKS = [
    ("A", "估值 · 殖利率", [("PE_TTM", "pe_ttm", MULT), ("PE Forward", "pe_ntm", MULT),
                        ("PEG_T", "peg_t", NUM2), ("PEG_F", "peg_f", NUM2),
                        ("EPS_TTM (Adj)", "eps_ttm", NUM2), ("EPS Forward", "eps_ntm", NUM2),
                        ("殖利率", "div_yield", PCT), ("每股配息 TTM", "dps_ttm", NUM2)]),
    ("D", "成長", [("營收QoQ 最新季", "rev_qoq1", PCT), ("  上一季 QoQ", "rev_qoq2", PCT),
                 ("  加速度 (pp)", "rev_accel", PP), ("  動能判讀", "momentum", None),
                 ("  同期加速度 (pp)", "rev_accel_sa", PP),
                 ("營收YoY 最新季", "rev_yoy1", PCT),
                 ("EPS_YoY 最新季", "eps_yoy1", PCT),
                 ("EPS成長_F (F1→F2)", "eps_g_f", PCT), ("當季EPS預估", "eps_q_est", NUM2)]),
    ("G", "毛利率 · ROE", [("毛利率_TTM", "gm_ttm", PCT), ("  5 年平均", "gm_5y", PCT),
                        ("  差距 (bps)", "gm_spread", BPS), ("  ΔGM_YoY (bps)", "gm_dyoy", BPS),
                        ("ROE_TTM", "roe_ttm", PCT), ("  5 年平均", "roe_5avg", PCT),
                        ("ROE Forward", "roe_f", PCT), ("每股淨值", "bvps", NUM2)]),
]
for colL, title, items in BLOCKS:
    ci = cif(colL)
    t = ws.cell(row=6, column=ci, value=title)
    t.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    t.fill = H_FILL; t.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=6, start_column=ci, end_row=6, end_column=ci + 1)
    for j, (lab, name, fmt) in enumerate(items):
        rr = 7 + j
        ws.cell(row=rr, column=ci, value=lab).font = Font(name=FONT, size=10)
        c = ws.cell(row=rr, column=ci + 1, value=CK(name))
        c.font = LNK_FONT; c.alignment = Alignment(horizontal="right")
        if fmt: c.number_format = fmt

ws["Z1"] = (f'=IF({CLK("n_an")}<Config!$B$7,"⚠ 分析師家數不足，共識預估可信度低　　","")'
            f'&IF(NOT(ISNUMBER({CLK("roe_ttm")})),"⚠ 股東權益為負，ROE 無比較意義　　","")'
            f'&IF({CLK("dps_ttm")}=0,"ℹ 此檔近四季未配息　　","")')
ws["A16"] = "旗標"; ws["A16"].font = SEC_FONT
ws["B16"] = '=IF($B$2="","",IF(LEN($Z$1)=0,"✓ 目前沒有觸發任何旗標",$Z$1))'
ws["B16"].font = Font(name=FONT, size=10, color="C00000")
ws.merge_cells("B16:L16")
ws.column_dimensions["Z"].hidden = True
ws.freeze_panes = "A3"

# ===================== 年度檢視 =====================
def AYS(val, r):
    return (f'SUMIFS({RQ(val)},{RQ("ticker")},{TSEL},'
            f'{RQ("fy")},$A{r},{RQ("is_est")},"N")')

def AQ4(val, fyexpr):
    return (f'SUMIFS({RQ(val)},{RQ("ticker")},{TSEL},'
            f'{RQ("fy")},{fyexpr},{RQ("fq")},4,{RQ("is_est")},"N")')

ws = wb.create_sheet("年度檢視")
ws.sheet_view.showGridLines = False
ws["A1"] = "年度檢視  ·  近六財年逐年指標"; ws["A1"].font = TITLE_FONT
ws["A2"] = "股票"; ws["A2"].font = Font(name=FONT, bold=True, size=10)
ws["B2"] = f'={TSEL}'; ws["B2"].font = Font(name=FONT, bold=True, size=12, color="008000")
ws["C2"] = ("← 在 Stock_Card 換股票，這一頁與「季度檢視」同步。每一列＝一個完整財年，"
            "全部由 Raw_Q 的季度資料 rollup 而來。")
ws["C2"].font = NOTE_FONT
ws["Z1"] = f'=IF({TSEL}="","",IFERROR(SUMIFS({RQ("fy")},{RQ("ticker")},{TSEL},{RQ("seq")},0),""))'
ws.column_dimensions["Z"].hidden = True

AH = [("財年", 9, "0"), ("實際季數", 9, "0"), ("營收($M)", 13, NUM0), ("營收YoY", 10, PCT),
      ("毛利率", 10, PCT), ("毛利率變化(bps)", 13, BPS), ("EPS(Adj)", 10, NUM2),
      ("EPS YoY", 10, PCT), ("ROE", 10, PCT), ("每股配息", 10, NUM2),
      ("年末股價", 11, NUM2), ("殖利率", 10, PCT), ("PE", 9, MULT), ("PEG_T", 9, NUM2)]
for j, (h, w, _) in enumerate(AH):
    ws.cell(row=4, column=1 + j, value=h)
    ws.column_dimensions[gcl(1 + j)].width = w
style_header(ws, 4, len(AH))

R0 = 5
for i in range(6):
    r = R0 + i
    ok = f'$B{r}>=4'
    ws.cell(row=r, column=1, value=f'=IF({TSEL}="","",IFERROR($Z$1-{5-i},""))').number_format = "0"
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=2,
            value=f'=IF({TSEL}="","",COUNTIFS({RQ("ticker")},{TSEL},'
                  f'{RQ("fy")},$A{r},{RQ("is_est")},"N"))').number_format = "0"
    ws.cell(row=r, column=3, value=f'=IF({ok},{AYS("revenue", r)},"")').number_format = NUM0
    ws.cell(row=r, column=4,
            value=f'=IFERROR(IF(AND(ISNUMBER(C{r}),ISNUMBER(C{r-1})),C{r}/C{r-1}-1,""),"")').number_format = PCT
    ws.cell(row=r, column=5,
            value=f'=IF({ok},IFERROR({AYS("gross_profit", r)}/C{r},""),"")').number_format = PCT
    ws.cell(row=r, column=6,
            value=f'=IFERROR(IF(AND(ISNUMBER(E{r}),ISNUMBER(E{r-1})),(E{r}-E{r-1})*10000,""),"")').number_format = BPS
    ws.cell(row=r, column=7, value=f'=IF({ok},{AYS("eps_diluted_adj", r)},"")').number_format = NUM2
    ws.cell(row=r, column=8,
            value=f'=IFERROR(IF(AND(ISNUMBER(G{r}),ISNUMBER(G{r-1})),G{r}/G{r-1}-1,""),"")').number_format = PCT
    e_now, e_prev = AQ4("total_equity", f"$A{r}"), AQ4("total_equity", f"$A{r}-1")
    ws.cell(row=r, column=9,
            value=f'=IF({ok},IFERROR(IF(MIN({e_now},{e_prev})<=0,"N/M",'
                  f'{NI_FY(f"$A{r}", TSEL)}/(({e_now}+{e_prev})/2)),"N/M"),"")').number_format = PCT
    ws.cell(row=r, column=10, value=f'=IF({ok},{AYS("dps", r)},"")').number_format = NUM2
    ws.cell(row=r, column=11,
            value=f'=IFERROR(IF({AQ4("price_at_end", f"$A{r}")}=0,"",'
                  f'{AQ4("price_at_end", f"$A{r}")}),"")').number_format = NUM2
    ws.cell(row=r, column=12,
            value=f'=IFERROR(IF(OR(NOT(ISNUMBER(K{r})),NOT(ISNUMBER(J{r})),J{r}=0),"",'
                  f'J{r}/K{r}),"")').number_format = PCT
    ws.cell(row=r, column=13,
            value=f'=IFERROR(IF(OR(NOT(ISNUMBER(K{r})),NOT(ISNUMBER(G{r})),G{r}<=0),"",'
                  f'K{r}/G{r}),"")').number_format = MULT
    ws.cell(row=r, column=14,
            value=f'=IFERROR(IF(OR(NOT(ISNUMBER(M{r})),NOT(ISNUMBER(H{r}))),"",'
                  f'IF(H{r}<Config!$B$6,"N/M",M{r}/(H{r}*100))),"")').number_format = NUM2
    for c in range(2, 15):
        ws.cell(row=r, column=c).font = FX_FONT

NC = len(AH)
for i, (lab, mapping) in enumerate([
    ("TTM / 最新季", dict(C="rev_ttm", D="rev_yoy1", E="gm_ttm", F="gm_spread",
                       G="eps_ttm", H="eps_ttm_yoy", I="roe_ttm", J="dps_ttm",
                       K="price", L="div_yield", M="pe_ttm", N="peg_t")),
    ("Forward（分析師共識）", dict(G="eps_ntm", H="eps_g_f", I="roe_f", M="pe_ntm", N="peg_f")),
]):
    r = 12 + i
    c0 = ws.cell(row=r, column=1, value=lab); c0.font = SEC_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for colL, name in mapping.items():
        d = next(x for x in CALC if x["name"] == name)
        cc = ws[f"{colL}{r}"]
        cc.value = (f'=IF({TSEL}="","",IFERROR(INDEX(Calc!${CL[name]}$2:${CL[name]}$100,'
                    f'MATCH({TSEL},Calc!$A$2:$A$100,0)),""))')
        cc.font = LNK_FONT
        cc.number_format = d["fmt"] or "General"
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="EDEDED")
ws.cell(row=11, column=1,
        value="↓ 以下兩列是「現況」，不是年度值。F 欄在 TTM 列是「最新毛利率 − 5 年平均」。").font = NOTE_FONT

ws["A15"] = "近五年摘要"; ws["A15"].font = SEC_FONT
SUMM = [("5 年平均毛利率", f'AVERAGE(E{R0}:E{R0+4})', PCT, "五個完整財年的算術平均"),
        ("最新毛利率 − 5 年平均 (bps)", f'(E12-AVERAGE(E{R0}:E{R0+4}))*10000', BPS,
         "正值＝目前毛利率高於五年常態"),
        ("5 年平均 ROE", f'AVERAGE(I{R0}:I{R0+4})', PCT, "N/M 的年度會被自動略過"),
        ("5 年平均殖利率", f'AVERAGE(L{R0}:L{R0+4})', PCT, "以各財年年末股價計算")]
for i, (lab, f, fmt, note) in enumerate(SUMM):
    r = 16 + i
    ws.cell(row=r, column=1, value=lab).font = Font(name=FONT, size=10)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=3, value=f'=IF({TSEL}="","",IFERROR({f},""))')
    c.font = FX_FONT; c.alignment = Alignment(horizontal="right")
    if fmt: c.number_format = fmt
    n = ws.cell(row=r, column=4, value=note); n.font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)

ws.conditional_formatting.add(f"F{R0}:F{R0+5}",
    ColorScaleRule(start_type="num", start_value=-200, start_color=RED,
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="num", end_value=200, end_color=GRN))
for _c in ("D", "L"):
    ws.conditional_formatting.add(f"{_c}{R0}:{_c}{R0+5}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color=GRN))
ws.conditional_formatting.add(f"M{R0}:M{R0+5}",
    ColorScaleRule(start_type="min", start_color=GRN, end_type="max", end_color=RED))
ws.freeze_panes = "C5"
ws.cell(row=22, column=1,
        value="PE 與殖利率都用「該財年年末股價」計算，不是用今天的股價除歷史數字——"
              "後者沒有估值意義。這需要 Raw_Q 的 price_at_end 欄。").font = NOTE_FONT
ws.cell(row=23, column=1,
        value="PEG_T ＝ 當年 PE ÷ 當年 EPS 成長率，是「市場當時為已實現成長付了多少」；"
              "PEG_F ＝ 前瞻 PE ÷ 預估成長率，才是估值指標。兩者不要混著解讀。").font = NOTE_FONT
ws.cell(row=24, column=1,
        value="年度層刻意不放加速度——加速度是季度層的訊號（這一季有沒有比上一季快），"
              "放在年度層會遲鈍到失去意義。").font = NOTE_FONT

# ===================== 季度檢視 =====================
ws = wb.create_sheet("季度檢視")
ws.sheet_view.showGridLines = False
ws["A1"] = "季度檢視  ·  近 12 季 + 當季預估"; ws["A1"].font = TITLE_FONT
ws["A2"] = "股票"; ws["A2"].font = Font(name=FONT, bold=True, size=10)
ws["B2"] = f'={TSEL}'; ws["B2"].font = Font(name=FONT, bold=True, size=12, color="008000")
ws["C2"] = "← 與 Stock_Card / 年度檢視 同步。每一列＝一個財季。加速度只做在這一層。"
ws["C2"].font = NOTE_FONT

def SQ(val, k):
    return f'SUMIFS({RQ(val)},{RQ("ticker")},{TSEL},{RQ("seq")},{k})'

def YOY(val, k):
    return f'({SQ(val, k)}/{SQ(val, k+4)}-1)'


def QOQ(val, k):
    """季度成長率：本季 ÷ 上一季 − 1"""
    return f'({SQ(val, k)}/{SQ(val, k+1)}-1)'

QH = [("期別", 12, None), ("期末日", 12, DATE), ("營收($M)", 13, NUM0),
      ("營收QoQ", 10, PCT), ("上一季QoQ", 11, PCT), ("加速度(pp)", 11, PP),
      ("動能判讀", 13, None), ("同期加速度(pp)", 13, PP), ("營收YoY", 10, PCT),
      ("毛利率", 10, PCT), ("ΔGM_YoY(bps)", 12, BPS), ("EPS(Adj)", 10, NUM2),
      ("EPS YoY", 10, PCT), ("每股配息", 10, NUM2), ("狀態", 9, None)]
for j, (h, w, _) in enumerate(QH):
    ws.cell(row=4, column=1 + j, value=h)
    ws.column_dimensions[gcl(1 + j)].width = w
style_header(ws, 4, len(QH))

QSEQ = list(range(12, -1, -1))
for i, k in enumerate(QSEQ):
    r = 5 + i
    g = lambda e: f'=IF({TSEL}="","",IFERROR({e},""))'
    ws.cell(row=r, column=1,
            value=f'=IF({TSEL}="","",IFERROR({QLOOK("period", k, 2, TSEL)},""))')
    ws.cell(row=r, column=2,
            value=g(f'IF({SQ("period_end", k)}=0,"",{SQ("period_end", k)})')).number_format = DATE
    ws.cell(row=r, column=3, value=g(SQ("revenue", k))).number_format = NUM0
    ws.cell(row=r, column=4, value=g(QOQ("revenue", k))).number_format = PCT
    ws.cell(row=r, column=5, value=g(QOQ("revenue", k + 1))).number_format = PCT
    ws.cell(row=r, column=6,
            value=g(f'({QOQ("revenue", k)}-{QOQ("revenue", k+1)})*100')).number_format = PP
    ws.cell(row=r, column=7,
            value=f'=IF({TSEL}="","",IFERROR(IF(NOT(ISNUMBER(D{r})),"",'
                  f'IF(D{r}>0,IF(F{r}>0,"成長且加速","成長但減速"),'
                  f'IF(F{r}>0,"衰退但收斂","衰退且惡化"))),""))')
    ws.cell(row=r, column=8,
            value=g(f'({QOQ("revenue", k)}-{QOQ("revenue", k+4)})*100')).number_format = PP
    ws.cell(row=r, column=9, value=g(YOY("revenue", k))).number_format = PCT
    ws.cell(row=r, column=10,
            value=g(f'{SQ("gross_profit", k)}/{SQ("revenue", k)}')).number_format = PCT
    ws.cell(row=r, column=11,
            value=g(f'({SQ("gross_profit", k)}/{SQ("revenue", k)}'
                    f'-{SQ("gross_profit", k+4)}/{SQ("revenue", k+4)})*10000')).number_format = BPS
    ws.cell(row=r, column=12, value=g(SQ("eps_diluted_adj", k))).number_format = NUM2
    ws.cell(row=r, column=13, value=g(YOY("eps_diluted_adj", k))).number_format = PCT
    ws.cell(row=r, column=14, value=g(SQ("dps", k))).number_format = NUM2
    ws.cell(row=r, column=15, value=f'=IF({TSEL}="","",IF({k}=0,"預估","實際"))')
    for c in range(1, 16):
        ws.cell(row=r, column=c).font = FX_FONT
LQ = 5 + len(QSEQ) - 1
ws.conditional_formatting.add(f"C5:C{LQ}",
    DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True))
ws.conditional_formatting.add(f"D5:D{LQ}",
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color=GRN))
for _txt, _fill, _col in [("成長且加速", "C6EFCE", "006100"), ("成長但減速", "FFEB9C", "9C5700"),
                          ("衰退但收斂", "FFEB9C", "9C5700"), ("衰退且惡化", "FFC7CE", "9C0006")]:
    ws.conditional_formatting.add(f"G5:G{LQ}",
        CellIsRule(operator="equal", formula=[f'"{_txt}"'],
                   font=Font(name=FONT, bold=True, color=_col),
                   fill=PatternFill("solid", fgColor=_fill)))
ws.conditional_formatting.add(f"H5:H{LQ}",
    ColorScaleRule(start_type="num", start_value=-8, start_color=RED,
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="num", end_value=8, end_color=GRN))
ws.conditional_formatting.add(f"F5:F{LQ}",
    ColorScaleRule(start_type="num", start_value=-30, start_color=RED,
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="num", end_value=30, end_color=GRN))
ws.conditional_formatting.add(f"K5:K{LQ}",
    ColorScaleRule(start_type="num", start_value=-200, start_color=RED,
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="num", end_value=200, end_color=GRN))
ws.conditional_formatting.add(f"A5:O{LQ}",
    FormulaRule(formula=['$O5="預估"'], fill=BOT_FILL,
                font=Font(name=FONT, italic=True, bold=True, color="BF8F00")))
ws.cell(row=LQ + 1, column=1,
        value="營收QoQ（季度成長率）＝ 本季營收 ÷ 上一季營收 − 1。"
              "加速度(pp) ＝ 本季QoQ − 上一季QoQ。注意：加速度是「成長率的變化」，不是營收的變化——"
              "營收在衰退但衰退幅度縮小時，加速度一樣是正的。「動能判讀」欄就是為了讓正負號不再誤導。").font = NOTE_FONT
ws.cell(row=LQ + 2, column=1,
        value="⚠ 加速度(pp) 會被季節性主宰：旺季那一季 QoQ 天生就高、淡季天生就低，"
              "所以這一欄每年會在同一季出現同樣的大幅震盪，那是季節性不是動能。"
              "同期加速度(pp) ＝ 本季QoQ − 去年同一財季的QoQ：一樣是在比「季度成長率」，"
              "但拿去年同一季當基準，季節性自動抵消，剩下的才是真正的動能變化。"
              "最後一列黃底＝當季預估。").font = NOTE_FONT

# --- 年 × 季 矩陣 ---
ws["Z1"] = f'=IF({TSEL}="","",IFERROR(SUMIFS({RQ("fy")},{RQ("ticker")},{TSEL},{RQ("seq")},0),""))'
ws.column_dimensions["Z"].hidden = True
for title, valcol, fmt, top, tot in [
        ("EPS (Adjusted)  ·  年 × 季矩陣", "eps_diluted_adj", NUM2, 22, True),
        ("營收 YoY 成長率  ·  年 × 季矩陣", None, PCT, 31, False)]:
    ws.cell(row=top, column=1, value=title).font = SEC_FONT
    heads = ["財年", "Q1", "Q2", "Q3", "Q4"] + (["FY 合計", "FY YoY"] if tot else [])
    for j, h in enumerate(heads):
        ws.cell(row=top + 1, column=1 + j, value=h)
    style_header(ws, top + 1, len(heads))
    for i in range(5):
        r = top + 2 + i
        ws.cell(row=r, column=1,
                value=f'=IF({TSEL}="","",IFERROR($Z$1-{4-i},""))').number_format = "0"
        ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10)
        for q in range(1, 5):
            cnt = f'COUNTIFS({RQ("ticker")},{TSEL},{RQ("fy")},$A{r},{RQ("fq")},{q})'
            if valcol:
                expr = (f'SUMIFS({RQ(valcol)},{RQ("ticker")},{TSEL},'
                        f'{RQ("fy")},$A{r},{RQ("fq")},{q})')
            else:
                cur = (f'SUMIFS({RQ("revenue")},{RQ("ticker")},{TSEL},'
                       f'{RQ("fy")},$A{r},{RQ("fq")},{q})')
                prv = (f'SUMIFS({RQ("revenue")},{RQ("ticker")},{TSEL},'
                       f'{RQ("fy")},$A{r}-1,{RQ("fq")},{q})')
                expr = f'IFERROR({cur}/{prv}-1,"")'
            c = ws.cell(row=r, column=1 + q, value=f'=IF({TSEL}="","",IF({cnt}=0,"",{expr}))')
            c.number_format = fmt; c.font = FX_FONT
            c.alignment = Alignment(horizontal="right")
        if tot:
            cf = f'COUNTIFS({RQ("ticker")},{TSEL},{RQ("fy")},$A{r})'
            c = ws.cell(row=r, column=6,
                        value=f'=IF({TSEL}="","",IF({cf}<4,"未滿4季",SUM(B{r}:E{r})))')
            c.number_format = fmt; c.font = Font(name=FONT, bold=True, size=10)
            if i > 0:
                c2 = ws.cell(row=r, column=7,
                             value=f'=IFERROR(IF(OR(NOT(ISNUMBER(F{r})),NOT(ISNUMBER(F{r-1}))),'
                                   f'"",F{r}/F{r-1}-1),"")')
                c2.number_format = PCT; c2.font = FX_FONT
    for q in range(1, 5):
        colL = gcl(1 + q)
        ws.conditional_formatting.add(f"{colL}{top+2}:{colL}{top+6}", FormulaRule(
            formula=[f'COUNTIFS({RQ("ticker")},{TSEL},{RQ("fy")},$A{top+2},'
                     f'{RQ("fq")},{q},{RQ("is_est")},"Y")>0'],
            fill=BOT_FILL, font=Font(name=FONT, italic=True, bold=True, color="BF8F00")))
    if not tot:
        ws.conditional_formatting.add(f"B{top+2}:E{top+6}",
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color=GRN))
ws.cell(row=21, column=1,
        value="以下兩張矩陣同一欄＝同一個財季。長條排列會把旺季誤讀成成長，"
              "排成矩陣才看得出季節性與真實趨勢。黃底斜體＝當季預估。").font = NOTE_FONT
ws.cell(row=30, column=1, value="↓ 這張就是「每季營收成長率」，季節可比。").font = NOTE_FONT
ws.freeze_panes = "B5"

# ===================== _Manifest =====================
ws = wb.create_sheet("_Manifest")
ws.sheet_view.showGridLines = False
ws["A1"] = "資料完整度儀表板"; ws["A1"].font = TITLE_FONT
MH = ["ticker", "實際季數", "目標", "最舊季", "最新實際季", "當季預估期別",
      "下次財報日", "距財報(天)", "資料日期", "狀態"]
for j, h in enumerate(MH):
    ws.cell(row=3, column=1 + j, value=h)
style_header(ws, 3, len(MH))
widths(ws, {"A": 10, "B": 10, "C": 8, "D": 12, "E": 13, "F": 14,
            "G": 12, "H": 11, "I": 12, "J": 18})
for i in range(NROW - 1):
    r, src = 4 + i, 2 + i
    ws.cell(row=r, column=1, value=f'=IF(Calc!$A${src}="","",Calc!$A${src})').font = LNK_FONT
    ws.cell(row=r, column=2,
            value=f'=IF($A{r}="","",COUNTIFS({RQ("ticker")},$A{r},{RQ("is_est")},"N"))'
            ).number_format = "0"
    ws.cell(row=r, column=3, value=f'=IF($A{r}="","",Config!$B$4)').number_format = "0"
    ws.cell(row=r, column=4,
            value=f'=IF($A{r}="","",IFERROR({SF("period_end", "$B" + str(r), r=r)},""))'
            ).number_format = DATE
    ws.cell(row=r, column=5,
            value=f'=IF($A{r}="","",IFERROR({QQ("period_end", 1, r)},""))').number_format = DATE
    ws.cell(row=r, column=6, value=f'=IF($A{r}="","",IFERROR({QLOOK("period", 0, r)},""))')
    ws.cell(row=r, column=7,
            value=f'=IF($A{r}="","",IFERROR({LOOK("Raw_Price","D",r)},""))').number_format = DATE
    ws.cell(row=r, column=8,
            value=f'=IF($A{r}="","",IFERROR({LOOK("Raw_Price","D",r)}-TODAY(),""))').number_format = "0"
    ws.cell(row=r, column=9,
            value=f'=IF($A{r}="","",IFERROR({LOOK("Raw_Price","E",r)},""))').number_format = DATE
    ws.cell(row=r, column=10,
            value=f'=IF($A{r}="","",IF($B{r}=0,"未抓取",'
                  f'IF($B{r}<$C{r},"需補歷史 ("&($C{r}-$B{r})&" 季)",'
                  f'IF(TODAY()>$G{r},"待更新財報","OK"))))')
    for c in range(2, 11):
        ws.cell(row=r, column=c).font = FX_FONT
for txt, fill, color in [('"OK"', "C6EFCE", "006100"), ('"待更新財報"', "FFEB9C", "9C5700")]:
    ws.conditional_formatting.add(f"J4:J{NROW + 2}",
        CellIsRule(operator="equal", formula=[txt], font=Font(name=FONT, color=color),
                   fill=PatternFill("solid", fgColor=fill)))
ws.conditional_formatting.add(f"J4:J{NROW + 2}",
    FormulaRule(formula=['ISNUMBER(SEARCH("補歷史",$J4))'],
                font=Font(name=FONT, color="9C0006"),
                fill=PatternFill("solid", fgColor="FFC7CE")))
ws.conditional_formatting.add(f"H4:H{NROW + 2}",
    CellIsRule(operator="lessThan", formula=["7"], font=Font(name=FONT, bold=True, color="9C0006")))
ws.freeze_panes = "A4"
ws.cell(row=NROW + 4, column=1,
        value="狀態欄告訴資料管線該做什麼：需補歷史 → 只抓缺的那幾季；待更新財報 → 只抓最新 1~2 季；"
              "OK → 整檔跳過季度 API。這是增量更新省下絕大部分請求的關鍵。").font = NOTE_FONT

# ===================== 說明 =====================
ws = wb.create_sheet("說明")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 26, "B": 96})
ROWS = [
    ("T", "美股觀察表  ·  架構範本 v2"), ("", ""),
    ("W", "資料由 GitHub Actions 自動更新。Raw 分頁為管線寫入區，請勿手改。"),
    ("", ""),
    ("H", "v2 改了什麼"),
    ("", "Raw_Q 從 32 欄砍到 14 欄。只留下你指定的指標，以及計算這些指標所必需的原始欄位。"),
    ("", "刪掉的是：營益率、FCF、杜邦拆解、ROIC、淨負債、有效稅率、adj_gap、PE 百分位、2年CAGR、"),
    ("", "近4季平均加速度、EPS加速度、年度加速度 —— 這些都是我加的，不是你要的。"),
    ("", "新增：殖利率（需要 Raw_Q 的 dps 欄）。"),
    ("", ""),
    ("H", "Raw_Q 的 14 欄，每一欄都有用途"),
    ("", "ticker / period / fy / fq / period_end   → 鍵值與期別對齊"),
    ("", "is_est / est_source                      → 區分實際與預估，避免預估污染歷史"),
    ("", "eps_basis                                → EPS 口徑標記：street=市場口徑；gaap=四層來源"),
    ("", "                                            全部失敗時的降級，PE/PEG 會失真，須注意"),
    ("", "revenue                                  → 營收成長率、毛利率分母"),
    ("", "gross_profit                             → 毛利率分子"),
    ("", "eps_diluted_adj                          → EPS、PE、PEG、EPS成長率"),
    ("", "shares_diluted                           → 反推 ROE 淨利、ROE_Forward 的每股淨值"),
    ("", "total_equity                             → ROE 分母"),
    ("", "dps                                      → 殖利率"),
    ("", "price_at_end                             → 歷史 PE、PEG_T、歷史殖利率"),
    ("", ""),
    ("H", "抓不到的東西（誠實說明）"),
    ("", "Adjusted 淨利：沒有任何 API 提供。它不是損益表項目，只存在於公司財報的 non-GAAP 對帳附註。"),
    ("", "本表用「逐季 adjusted EPS × 稀釋股數」反推，逐季相乘再加總比用單一股數準，但仍是近似值。"),
    ("", "Adjusted EPS 本身拿得到，但不是從損益表，是從「歷史財報實際值」那類端點的 actual EPS。"),
    ("", "fy / fq：多數來源不直接提供公司自己的財年標號，要從 period_end 推算。"),
    ("", "歷史的 Forward PE（每一年當時的前瞻本益比）：拿不到，歷史分析師共識一般不留存。"),
    ("", ""),
    ("H", "年 與 季 分在哪裡"),
    ("", "年度層 → 「年度檢視」：每一列＝一個完整財年，由 Raw_Q 的季度資料 rollup 而來。"),
    ("", "季度層 → 「季度檢視」：每一列＝一個財季，另有年 × 季矩陣做季節可比的閱讀。"),
    ("", "橫向層 → 「Dashboard」：每一列＝一檔股票，用來跨標的比較與篩選。"),
    ("", "三層共用同一份 Raw_Q，年度值永遠等於該年四季加總，不會有兩套打架的資料。"),
    ("", ""),
    ("H", "加速度只做在季度層"),
    ("", "季度成長率(QoQ) ＝ 本季營收 ÷ 上一季營收 − 1。"),
    ("", "加速度(pp) ＝ 本季QoQ − 上一季QoQ。字面上就是「這季成長率有沒有比上季快」。"),
    ("", "但這一欄會被季節性主宰：旺季 QoQ 天生高、淡季天生低，每年在同一季重複同樣的震盪。"),
    ("", "同期加速度(pp) ＝ 本季QoQ − 去年同一財季的QoQ。一樣在比季度成長率，但基準換成"),
    ("", "去年同一季，季節性自動抵消。實務上這一欄才有訊號，建議以它為主、加速度當參考。"),
    ("", "年度層刻意不放加速度：一年才一個點，訊號遲鈍到失去意義。"),
    ("", ""),
    ("H", "口徑"),
    ("", "EPS / PE / PEG / EPS成長 → Adjusted，與分析師預估和市場定價同口徑。"),
    ("", "ROE → 分子用 adjusted EPS 反推的淨利，分母用 GAAP 股東權益（你選的口徑）。"),
    ("", "注意：這代表 ROE 數字會與各大財經網站（多用 GAAP 淨利）不同，這是刻意的。"),
    ("", ""),
    ("H", "顏色慣例"),
    ("", "藍字＝你手動輸入 ｜ 黑字＝公式 ｜ 綠字＝跨分頁連結 ｜ 黃底＝資料管線寫入區，請勿手改"),
    ("", ""),
    ("H", "新增一檔股票"),
    ("", "① Tickers 第 2~21 列往下打一列 → ② 重跑資料管線 → ③ Excel 按「資料 → 全部重新整理」"),
]
r = 1
for kind, text in ROWS:
    c = ws.cell(row=r, column=1 if kind in ("T", "H") else 2, value=text)
    if kind == "T":
        c.font = TITLE_FONT
    elif kind == "H":
        c.font = SEC_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    elif kind == "W":
        c.font = WARN_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
    else:
        c.font = Font(name=FONT, size=10)
    r += 1

# ===================== 存檔 =====================
wb["Calc"].sheet_state = "hidden"
wb._sheets = [wb[n] for n in ["說明", "Dashboard", "Stock_Card", "年度檢視", "季度檢視",
                              "_Manifest", "Tickers", "Raw_Q", "Raw_Est", "Raw_Price",
                              "Calc", "Config"]]
wb.active = 1
OUT = Path(__file__).parent / "美股觀察表.xlsx"
wb.save(OUT)
print(f"saved {OUT.name}  ({len(TICKERS)} 檔 / {nq} 季度列)")
